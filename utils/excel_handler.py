"""
Excel Handler
Utilities for reading and writing Excel files.
"""

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from typing import Dict, List, Tuple, Optional, Any
from datetime import datetime
import os


class ExcelReader:
    """Handles reading Excel files and extracting sheet data"""
    
    SUPPORTED_EXTENSIONS = ['.xlsx', '.xls']
    
    @staticmethod
    def is_supported(filepath: str) -> bool:
        """Check if file extension is supported"""
        ext = os.path.splitext(filepath)[1].lower()
        return ext in ExcelReader.SUPPORTED_EXTENSIONS
    
    @staticmethod
    def get_sheet_names(filepath: str) -> List[str]:
        """Get list of sheet names from Excel file"""
        try:
            xl = pd.ExcelFile(filepath)
            return xl.sheet_names
        except PermissionError:
            raise ValueError(f"Permission denied: Cannot access '{filepath}'. The file may be open in another program.")
        except FileNotFoundError:
            raise ValueError(f"File not found: '{filepath}'")
        except Exception as e:
            raise ValueError(f"Error reading file '{filepath}': {str(e)}")
    
    @staticmethod
    def read_sheet(filepath: str, sheet_name: str) -> pd.DataFrame:
        """Read a specific sheet from Excel file"""
        try:
            # Read without headers (we'll detect them later)
            df = pd.read_excel(filepath, sheet_name=sheet_name, header=None)
            return df
        except PermissionError:
            raise ValueError(f"Permission denied: Cannot access '{filepath}'. The file may be open in another program.")
        except FileNotFoundError:
            raise ValueError(f"File not found: '{filepath}'")
        except Exception as e:
            raise ValueError(f"Error reading sheet '{sheet_name}' from '{filepath}': {str(e)}")
    
    @staticmethod
    def read_multiple_sheets(filepath: str, sheet_names: List[str]) -> Dict[str, pd.DataFrame]:
        """Read multiple sheets from Excel file"""
        result = {}
        errors = []
        for sheet_name in sheet_names:
            try:
                df = ExcelReader.read_sheet(filepath, sheet_name)
                result[sheet_name] = df
            except Exception as e:
                result[sheet_name] = None
                errors.append(f"Sheet '{sheet_name}': {str(e)}")
        # Log errors if any sheets failed (callers can check for None values)
        if errors:
            import logging
            logging.warning(f"Failed to read sheets from {filepath}: {'; '.join(errors)}")
        return result
    
    @staticmethod
    def get_all_columns(sources: List[Tuple[str, str, pd.DataFrame]]) -> List[str]:
        """
        Get union of all columns from all sources.
        sources: List of (filepath, sheet_name, dataframe) tuples
        Returns unique column names.
        """
        all_columns = set()
        
        for filepath, sheet_name, df in sources:
            if df is not None and len(df) > 0:
                # Try to find header row and get column names
                try:
                    # Look for header row with Date and Particulars
                    for idx in range(min(20, len(df))):
                        row = df.iloc[idx]
                        first_three = [str(val).strip() if pd.notna(val) else '' for val in row.iloc[:3]]
                        if 'Date' in first_three and 'Particulars' in first_three:
                            headers = df.iloc[idx].tolist()
                            headers = [str(h).strip() if pd.notna(h) else f'Column_{i}' 
                                       for i, h in enumerate(headers)]
                            all_columns.update(headers)
                            break
                except Exception:
                    pass
        
        return sorted(list(all_columns))


class ExcelWriter:
    """Handles writing processed data to Excel files"""
    
    # Styling constants
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    TAX_COL_FILL = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")  # Light blue for Tax
    TAXABLE_COL_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green for Taxable
    EXCLUDE_COL_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Light yellow for Excluded
    STANDARD_COL_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # Light sage for Standard
    # Data sheet color coding (matching app colors)
    CGST_SGST_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # Light green for CGST/SGST
    IGST_FILL = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")  # Light blue for IGST
    TOTAL_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")  # Light orange for Totals
    TAXABLE_DATA_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # Light green for taxable data
    BORDER = Border(
        left=Side(style='thin', color='B4B4B4'),
        right=Side(style='thin', color='B4B4B4'),
        top=Side(style='thin', color='B4B4B4'),
        bottom=Side(style='thin', color='B4B4B4')
    )
    
    @staticmethod
    def write_output(
        filepath: str,
        gst_data: pd.DataFrame,
        non_gst_data: pd.DataFrame,
        metadata: Dict[str, Any],
        tax_config: List[Dict],
        exclusion_list: List[str],
        column_types: Dict[str, str],
        column_stats: Dict[str, Dict] = None
    ):
        """
        Write complete output to Excel file with 4 sheets.
        
        Args:
            filepath: Output file path
            gst_data: GST transactions DataFrame
            non_gst_data: Non-GST transactions DataFrame
            metadata: Processing metadata
            tax_config: Tax configuration list
            exclusion_list: Exclusion column list
            column_types: Dict mapping column names to their types
            column_stats: Dict mapping column names to their stats (type, sum)
        """
        from openpyxl.worksheet.table import Table, TableStyleInfo
        
        wb = Workbook()
        
        # Sheet 1: Metadata
        ws_meta = wb.active
        ws_meta.title = "Metadata"
        ExcelWriter._write_metadata_sheet(ws_meta, metadata, column_types, column_stats)
        
        # Sheet 2: Configuration
        ws_config = wb.create_sheet("Configuration")
        ExcelWriter._write_config_sheet(ws_config, tax_config, exclusion_list)
        
        # Sheet 3: GST Transactions (as Excel Table with index)
        ws_gst = wb.create_sheet("GST Transactions")
        ExcelWriter._write_data_sheet_as_table(ws_gst, gst_data, "GST", "GSTTable", column_types)

        # Sheet 4: Non-GST Transactions (as Excel Table with index)
        ws_non_gst = wb.create_sheet("Non-GST Transactions")
        ExcelWriter._write_data_sheet_as_table(ws_non_gst, non_gst_data, "Non-GST", "NonGSTTable", column_types)
        
        # Save
        wb.save(filepath)
    
    @staticmethod
    def _write_metadata_sheet(ws, metadata: Dict[str, Any], column_types: Dict[str, str], column_stats: Dict[str, Dict] = None):
        """Write metadata sheet"""
        # Title
        ws['A1'] = "GST Compliance Tool - Processing Report"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:D1')
        
        # Processing info
        ws['A3'] = "Processing Date:"
        ws['B3'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        ws['A4'] = "Total Sources:"
        ws['B4'] = metadata.get('total_sources', 0)
        
        ws['A5'] = "Total GST Rows:"
        ws['B5'] = metadata.get('total_gst_rows', 0)
        
        ws['A6'] = "Total Non-GST Rows:"
        ws['B6'] = metadata.get('total_non_gst_rows', 0)
        
        # Source files table
        ws['A8'] = "Source Files Summary"
        ws['A8'].font = Font(bold=True, size=12)

        headers = ['Source File', 'Sheet', 'Original Rows', 'Data Rows', 'GST Rows', 'Non-GST Rows', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=9, column=col, value=header)
            cell.fill = ExcelWriter.HEADER_FILL
            cell.font = ExcelWriter.HEADER_FONT
            cell.border = ExcelWriter.BORDER

        row = 10
        sources = metadata.get('sources', [])
        for source in sources:
            source_name = source.get('source', '')
            parts = source_name.split(' | ')
            file_name = parts[0] if parts else source_name
            sheet_name = parts[1] if len(parts) > 1 else ''

            ws.cell(row=row, column=1, value=file_name).border = ExcelWriter.BORDER
            ws.cell(row=row, column=2, value=sheet_name).border = ExcelWriter.BORDER
            ws.cell(row=row, column=3, value=source.get('original_rows', 0)).border = ExcelWriter.BORDER
            # Data Rows = rows after cleaning (header/grand total removed)
            data_rows = source.get('rows_after_cleaning', source.get('gst_rows', 0) + source.get('non_gst_rows', 0))
            ws.cell(row=row, column=4, value=data_rows).border = ExcelWriter.BORDER
            ws.cell(row=row, column=5, value=source.get('gst_rows', 0)).border = ExcelWriter.BORDER
            ws.cell(row=row, column=6, value=source.get('non_gst_rows', 0)).border = ExcelWriter.BORDER

            error = source.get('error')
            status = 'Error: ' + error if error else 'Success'
            ws.cell(row=row, column=7, value=status).border = ExcelWriter.BORDER

            row += 1

        # Totals row
        ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=row, column=3, value=sum(s.get('original_rows', 0) for s in sources)).font = Font(bold=True)
        total_data_rows = sum(s.get('rows_after_cleaning', s.get('gst_rows', 0) + s.get('non_gst_rows', 0)) for s in sources)
        ws.cell(row=row, column=4, value=total_data_rows).font = Font(bold=True)
        ws.cell(row=row, column=5, value=metadata.get('total_gst_rows', 0)).font = Font(bold=True)
        ws.cell(row=row, column=6, value=metadata.get('total_non_gst_rows', 0)).font = Font(bold=True)
        
        # Column Configuration table with Index, Type, Sum
        row += 3
        ws.cell(row=row, column=1, value="Column Configuration").font = Font(bold=True, size=12)
        row += 1
        
        col_headers = ['#', 'Column Name', 'Type', 'Sum']
        for col, header in enumerate(col_headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = ExcelWriter.HEADER_FILL
            cell.font = ExcelWriter.HEADER_FONT
            cell.border = ExcelWriter.BORDER
        
        row += 1
        for idx, (col_name, col_type) in enumerate(sorted(column_types.items()), 1):
            # Index column
            cell_idx = ws.cell(row=row, column=1, value=idx)
            cell_idx.border = ExcelWriter.BORDER
            cell_idx.alignment = Alignment(horizontal='center')
            
            # Column name
            cell1 = ws.cell(row=row, column=2, value=col_name)
            cell1.border = ExcelWriter.BORDER
            
            # Type
            cell2 = ws.cell(row=row, column=3, value=col_type)
            cell2.border = ExcelWriter.BORDER
            
            # Sum (from column_stats if available)
            sum_value = ""
            if column_stats and col_name in column_stats:
                stats = column_stats[col_name]
                if stats.get('is_numeric', False):
                    sum_value = stats.get('sum', 0)
                else:
                    sum_value = "Text"
            
            cell3 = ws.cell(row=row, column=4, value=sum_value)
            cell3.border = ExcelWriter.BORDER
            if isinstance(sum_value, (int, float)):
                cell3.number_format = '#,##0.00'
            
            # Color code based on type - different colors for each type
            if col_type == 'Tax':
                fill = ExcelWriter.TAX_COL_FILL  # Light blue
            elif col_type == 'Taxable':
                fill = ExcelWriter.TAXABLE_COL_FILL  # Light green
            elif col_type == 'Excluded':
                fill = ExcelWriter.EXCLUDE_COL_FILL  # Light yellow
            elif col_type == 'Standard':
                fill = ExcelWriter.STANDARD_COL_FILL  # Light sage
            else:
                fill = None
            
            if fill:
                cell_idx.fill = fill
                cell1.fill = fill
                cell2.fill = fill
                cell3.fill = fill
            
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 20
    
    @staticmethod
    def _write_config_sheet(ws, tax_config: List[Dict], exclusion_list: List[str]):
        """Write configuration sheet"""
        # Tax Config table
        ws['A1'] = "Tax Configuration"
        ws['A1'].font = Font(bold=True, size=12)
        
        tax_headers = ['TaxType', 'TaxRate', 'ColumnNames', 'Delimiter']
        for col, header in enumerate(tax_headers, 1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.fill = ExcelWriter.HEADER_FILL
            cell.font = ExcelWriter.HEADER_FONT
            cell.border = ExcelWriter.BORDER
        
        for row_idx, config_row in enumerate(tax_config, 3):
            for col_idx, header in enumerate(tax_headers, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=config_row.get(header, ''))
                cell.border = ExcelWriter.BORDER
        
        # Exclusion List table (to the right)
        ws['F1'] = "Exclusion List"
        ws['F1'].font = Font(bold=True, size=12)
        
        cell = ws.cell(row=2, column=6, value="ExcludeColumn")
        cell.fill = ExcelWriter.HEADER_FILL
        cell.font = ExcelWriter.HEADER_FONT
        cell.border = ExcelWriter.BORDER
        
        for row_idx, excl_col in enumerate(exclusion_list, 3):
            cell = ws.cell(row=row_idx, column=6, value=excl_col)
            cell.border = ExcelWriter.BORDER
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['F'].width = 20
    
    @staticmethod
    def _write_data_sheet_as_table(ws, df: pd.DataFrame, sheet_type: str, table_name: str,
                                    column_types: Dict[str, str] = None):
        """Write data sheet as Excel Table (Ctrl+T style) with index column and highlighted columns"""
        from openpyxl.worksheet.table import Table, TableStyleInfo

        if df is None or len(df) == 0:
            ws['A1'] = f"No {sheet_type} transactions found"
            return
        
        # Add index column to dataframe (avoid collision with existing '#' column)
        df_with_index = df.copy()

        # Deduplicate column names to prevent Excel table corruption
        cols = df_with_index.columns.tolist()
        seen = {}
        new_cols = []
        for col in cols:
            col_str = str(col) if col is not None else 'Column'
            if col_str in seen:
                seen[col_str] += 1
                new_cols.append(f"{col_str}_{seen[col_str]}")
            else:
                seen[col_str] = 0
                new_cols.append(col_str)
        df_with_index.columns = new_cols

        index_col_name = '#'
        if index_col_name in df_with_index.columns:
            # Find a unique name for index column
            counter = 1
            while f'{index_col_name}{counter}' in df_with_index.columns:
                counter += 1
            index_col_name = f'#{counter}'
        df_with_index.insert(0, index_col_name, range(1, len(df) + 1))
        
        # Define computed/derived columns
        computed_columns = {
            'Active Columns', 'Taxable Value', 'Tax Rates (Config)', 'Tax Rates (Calculated)',
            'Transaction Type', 'Review Required', 'Source'
        }

        # Determine fill color for each column based on type
        def get_column_fill(col_name):
            """Get the appropriate fill color for a column based on its type"""
            if col_name.startswith(('CGST_', 'SGST_')):
                return ExcelWriter.CGST_SGST_FILL  # Light green
            elif col_name.startswith('IGST_'):
                return ExcelWriter.IGST_FILL  # Light blue
            elif col_name.startswith('Total '):
                return ExcelWriter.TOTAL_FILL  # Light orange
            elif col_name == 'CESS':
                return ExcelWriter.CGST_SGST_FILL  # Light green (same as CGST/SGST)
            elif col_name in computed_columns:
                return ExcelWriter.TAXABLE_COL_FILL  # Light green for other computed
            # Highlight Taxable data columns (raw data used in taxable value calculation)
            elif column_types and col_name in column_types and column_types[col_name] == 'Taxable':
                return ExcelWriter.TAXABLE_DATA_FILL  # Light green for taxable data
            return None

        # Track column fills by index
        column_fills = {}
        for col_idx, col_name in enumerate(df_with_index.columns, 1):
            fill = get_column_fill(col_name)
            if fill:
                column_fills[col_idx] = fill
        
        # Write headers
        for col_idx, col_name in enumerate(df_with_index.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            # Headers will be styled by the table
        
        # Write data with highlighting
        for row_idx, row in enumerate(df_with_index.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)

                # Format numbers
                if isinstance(value, (int, float)) and not pd.isna(value):
                    if col_idx == 1:  # Index column - no decimals
                        cell.number_format = '#,##0'
                    else:
                        cell.number_format = '#,##0.00'

                # Apply column-specific fill color
                if col_idx in column_fills:
                    cell.fill = column_fills[col_idx]
        
        # Auto-adjust column widths
        for col_idx, col_name in enumerate(df_with_index.columns, 1):
            max_length = len(str(col_name))
            for row_idx in range(2, min(102, len(df_with_index) + 2)):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
        
        # Create Excel Table
        last_col = get_column_letter(len(df_with_index.columns))
        last_row = len(df_with_index) + 1
        table_ref = f"A1:{last_col}{last_row}"
        
        # Create table with style
        table = Table(displayName=table_name, ref=table_ref)
        
        # Use a medium blue style (TableStyleMedium2 is a nice blue)
        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        table.tableStyleInfo = style
        
        ws.add_table(table)
        
        # NO freeze panes - user requested no freezing


class ExcelExporter:
    """Simplified exporter for quick exports"""
    
    @staticmethod
    def quick_export(df: pd.DataFrame, filepath: str, sheet_name: str = "Data"):
        """Quick export DataFrame to Excel"""
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
    
    @staticmethod
    def export_multiple_sheets(sheets: Dict[str, pd.DataFrame], filepath: str):
        """Export multiple DataFrames to different sheets"""
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            for sheet_name, df in sheets.items():
                if df is not None and len(df) > 0:
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
